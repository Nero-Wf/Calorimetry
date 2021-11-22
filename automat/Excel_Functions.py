# library/modules from python:
from openpyxl import Workbook

# own scripts:
import automat.Dictionary as Dictionary

def create_excel(substance_data, file_name):   
    wb = Workbook()

    # create all sheets
    sheet = []
    sheet_names = ["Evaluation", "Raw_Data_COM"]

    sheet.append(wb.active)
    sheet[0].title = sheet_names[0]
    sheet.append(wb.create_sheet(sheet_names[1]))

    # setup the evaluation sheet
    counter = 1
    ret_counter = []

    title = [["Substance Data"], ["Process setup"], ["Raw Data Processing"], ["Calculation"]]
    data = [["Substance", "Molar Mass [g/mol]", "Weighing [g]", "Volume [ml]", "Concentration [mol/l]"], 
            ["Process Points", "Evaluation Start Time", "Evaluation End Time", "V_A [ml/min]", "V_A,act [ml/min]", "n_A,act [mol/s]", "n_A,act,water [mol/s]", "V_B [ml/min]", "V_B,act [ml/min]", "n_B,act [mol/s]", "n_B,act,water [mol/s]"],
            ["Process Points", "T_A [°C]", "T_B [°C]", "T_out [°C]", "Upre [V]", "Ur1 [V]", "Ur2 [V]", "dT_A [°C]", "dT_B [°C]", "dT_out [°C]", "Q_Out [W]"],
            ["Process Points", "Q_A [W]", "Q_B [W]", "Qpre [W] - cp flux", "QSE,pre [W]", "Qr1 [W]", "Qr2 [W]", "dHr [kJ/mol]"]]

    for idx in range(4):
        sheet[0].append(title[idx])
        sheet[0].append(data[idx])
        counter += 2
        ret_counter.append([counter-2, counter, len(data[idx])])

    # insert substance data to the evaluation sheet
    sheet[0].insert_rows(idx=3, amount = 3)
    for idx in range(2):
        sheet[0].cell(row=idx+3, column=1).value = chr(idx+97)
        sheet[0].cell(row=idx+3, column=2).value = substance_data.get_molar_mass()[idx]
        sheet[0].cell(row=idx+3, column=3).value = substance_data.get_weighing()[idx]
        sheet[0].cell(row=idx+3, column=4).value = substance_data.get_volume()[idx]
        sheet[0].cell(row=idx+3, column=5).value = substance_data.get_concentration()[idx]

    sheet[0].cell(row=2, column=8).value = "Additional data"
    sheet[0].cell(row=3, column=8).value = "concentration [mol/l]"
    sheet[0].cell(row=3, column=9).value = Dictionary.calculation_data["concentration"]
    sheet[0].cell(row=4, column=8).value = "cp [J/(molK)]"
    sheet[0].cell(row=4, column=9).value = Dictionary.calculation_data["cp"]

    # remember at which position the titles are and where rows have to be inserted later 
    sheet[0].insert_rows(idx=8, amount=1)
    sheet[0].insert_rows(idx=11, amount=1)

    val = 3
    for idx in range(1,4):
        ret_counter[idx][0] += val
        ret_counter[idx][1] += val
        val += 1  
    ret_counter[0][1] = 5

    # setup the evaluation sheet
    sheet[1].append(["Elapsed_Time", "T_set", "T_pre", "T_r1", "T_r2", "T_A", "T_B", "T_out", "U_pre", "U_r1", "U_r2"])

    wb.save("{}.xlsx".format(file_name))

    return wb, sheet, ret_counter