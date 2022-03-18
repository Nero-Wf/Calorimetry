# library/modules from python:
import time
import math 
from enum import Enum
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, colors

# own scripts:
import automat.pyStrategy as pyStrategy
import automat.Excel_Functions as Excel_Functions
import automat.Dictionary as Dictionary

class operation_point_list_entry:
    """This class turns the user's input into an object, making it easier to handle the operating points."""
    def __init__(self, time_ms, temperature, flowrate_list):
        self.time_ms = time_ms
        self.temperature = round(temperature)
        self.flowrate_list = flowrate_list
    
    def get_time_ms(self):
        return self.time_ms

    def get_temperature(self):
        return self.temperature

    def get_flowrate(self, idx):
        return self.flowrate_list[idx]
    
    def get_number_of_pumps(self):
        return len(self.flowrate_list)

class substance_data:
    def __init__(self, weighing_g, volume_ml, molar_mass_gpermol, pump_substance_assignment_list):
        self.weighing = weighing_g
        self.volume = volume_ml 
        self.molar_mass = molar_mass_gpermol
        self.list = pump_substance_assignment_list

        if len(self.weighing) != 2 or len(self.volume) != 2 or len(self.molar_mass) != 2:
            raise Exception("Substance data is not complete")

    def get_weighing(self):
        return self.weighing 

    def get_volume(self):
        return self.volume

    def get_molar_mass(self):
        return self.molar_mass

    def get_concentration(self):
        concentration = [] # mol/l
        for idx in range(2):
            concentration.append(self.weighing[idx]/(self.volume[idx] * 1E-3)/self.molar_mass[idx])
        return concentration

    def get_pump_substance_assignment_list(self):
        return self.list

class Output_Calculation_Absolute_Evaluation(pyStrategy.Strategy_Base):
    class States(Enum):
        TEMPERATURE_EQUILIBRATION = 0,
        SETTING_DEADLINE = 1,
        WAITING_FOR_DEADLINE = 2,

    def __init__(self, operation_point_list, substance_data, dead_time_ms, excel_name):
        self.list = operation_point_list  
        self.substance_data = substance_data
        self.dead_time = dead_time_ms

        self.idx = 0
        self.cur_temp = float("NaN")
        self.cur_deadline = 0
        self.min_time = 0
        self.cur_operation_point = None
        self.state = Output_Calculation_Absolute_Evaluation.States.TEMPERATURE_EQUILIBRATION
        self.datalist = []
        
        # variables for calculation
        self.process_point = 0

        # create excel file
        self.excel_name = excel_name     
        [self.workbook, self.sheet, self.counter] = Excel_Functions.create_excel(self.substance_data, self.excel_name)

        # sanity check
        for idx in range(len(self.list)):
            if not len(self.substance_data.list) == len(self.list[idx].flowrate_list):
                raise Exception("Length of substance pump assignment list and flow rate list do not match")

            if not self.dead_time < self.list[idx].time_ms:
                raise Exception("The dead time is longer than the operating time, so there is no evaluation time")

            try:
                tmp = "{:d}".format(int(self.list[idx].temperature))
                Dictionary.calorimeter_thermostat[tmp]
            except KeyError:
                raise Exception("No calorimeter calibration is given for the given set temperature")

    def get_operation_point(self):
        if not self.idx < len(self.list):
            self.cur_operation_point = None
            return None

        self.cur_operation_point = self.list[self.idx]
        if not self.cur_operation_point.get_temperature() == self.cur_temp:
            self.cur_deadline = time.monotonic_ns() + 10 * 60 * 1E9
            self.state = Output_Calculation_Absolute_Evaluation.States.TEMPERATURE_EQUILIBRATION
            self.cur_temp =  self.cur_operation_point.get_temperature()
            return pyStrategy.Strategy_Base.operation_point_information(self.cur_temp, [0] * self.cur_operation_point.get_number_of_pumps())
        
        self.state = Output_Calculation_Absolute_Evaluation.States.SETTING_DEADLINE
        self.idx += 1
        return pyStrategy.Strategy_Base.operation_point_information(self.cur_temp, self.cur_operation_point.flowrate_list)

    def push_value(self, line):
        if line is None:
            return
        
        self.datalist.append(line)
        self.sheet[1].append(line)

        if not self.state == Output_Calculation_Absolute_Evaluation.States.WAITING_FOR_DEADLINE:
            return
   
        # one-time calculation
        if self.min_time < time.monotonic_ns() and self.waiting_counter == 0:
            self.starting_idx = len(self.datalist)-1
            self.waiting_counter = 1

            self.process_point += 1
            self.evalutaion_time = [self.datalist[self.starting_idx][0], None]
            self.set_volume_flowrate = [0, 0]
            self.actual_volume_flowrate = [0, 0]
            self.actual_molar_flowrate = [0, 0]
            self.actual_water_molar_flowrate = [0, 0]

            for idx in range(1,4):
                self.sheet[0].insert_rows(idx=self.counter[idx][1], amount=1)
                if idx == 1:
                    for jdx in range(1,4):
                        self.counter[jdx][1] += 1
                        if jdx != 1:
                            self.counter[jdx][0] += 1    
                if idx == 2:
                    for jdx in range(2,4):
                        self.counter[jdx][1] += 1
                    self.counter[3][0] += 1 
                if idx == 3:
                    self.counter[idx][1] += 1              

            for idx in range(len(self.substance_data.list)):
                if self.substance_data.list[idx] == "A":
                    jdx = 0
                if self.substance_data.list[idx] == "B":
                    jdx = 1
                self.set_volume_flowrate[jdx] += self.cur_operation_point.flowrate_list[idx]
                self.actual_volume_flowrate[jdx] += self.actual_flowrate_list[idx]
                self.actual_molar_flowrate[jdx] += self.actual_flowrate_list[idx] * self.substance_data.get_concentration()[idx] / 6E4
                self.actual_water_molar_flowrate[jdx] += self.actual_flowrate_list[idx] * Dictionary.calculation_data["concentration"] / 6E4

            # process setup entry    
            self.sheet[0].cell(row=self.counter[1][1]-1, column=1).value = self.process_point
            self.sheet[0].cell(row=self.counter[1][1]-1, column=2).value = self.evalutaion_time[0]
            self.sheet[0].cell(row=self.counter[1][1]-1, column=4).value = self.set_volume_flowrate[0]
            self.sheet[0].cell(row=self.counter[1][1]-1, column=5).value = self.actual_volume_flowrate[0]
            self.sheet[0].cell(row=self.counter[1][1]-1, column=6).value = self.actual_molar_flowrate[0]
            self.sheet[0].cell(row=self.counter[1][1]-1, column=7).value = self.actual_water_molar_flowrate[0]
            self.sheet[0].cell(row=self.counter[1][1]-1, column=8).value = self.set_volume_flowrate[1]
            self.sheet[0].cell(row=self.counter[1][1]-1, column=9).value = self.actual_volume_flowrate[1]
            self.sheet[0].cell(row=self.counter[1][1]-1, column=10).value = self.actual_molar_flowrate[1]
            self.sheet[0].cell(row=self.counter[1][1]-1, column=11).value = self.actual_water_molar_flowrate[1]
                            
        # ongoing calculation
        if self.waiting_counter == 1:
            mean_values = []
            temp_difference = []
            heat_flux_outside = []
            heat_flux_reactor = None
            enthalpy_difference = None

            # process setup entry
            self.evalutaion_time[1] = self.datalist[len(self.datalist)-1][0]
            self.sheet[0].cell(row=self.counter[1][1]-1, column=3).value = self.evalutaion_time[1]

            # raw data processing entry (mean values)
            self.sheet[0].cell(row=self.counter[2][1]-1, column=1).value = self.process_point
            for idx in range(5,11):
                mean = 0
                counter = 0
                for jdx in range(self.starting_idx, len(self.datalist)):
                    counter += 1
                    mean += self.datalist[jdx][idx]
                mean_values.append(mean/counter)
                self.sheet[0].cell(row=self.counter[2][1]-1, column=idx-3).value = mean_values[idx-5]
            
            # raw data processing and calculation entry (temperature difference and outside heat flux)
            for idx in range(3):
                temp_difference.append(self.cur_operation_point.temperature - mean_values[idx])
                self.sheet[0].cell(row=self.counter[2][1]-1, column=idx+8).value = temp_difference[idx]

                if not idx == 2:
                    tmp = self.actual_volume_flowrate[idx] * Dictionary.calculation_data["concentration"] * Dictionary.calculation_data["cp"] * temp_difference[idx] / 6E4
                    heat_flux_outside.append(tmp)
                    self.sheet[0].cell(row=self.counter[3][1]-1, column=idx+2).value = heat_flux_outside[idx]

                else:
                    tmp = sum(self.actual_water_molar_flowrate) * Dictionary.calculation_data["cp"] * temp_difference[idx]
                    heat_flux_outside.append(tmp)
                    self.sheet[0].cell(row=self.counter[2][1]-1, column=self.counter[2][2]).value = heat_flux_outside[idx]

            # calculation entry (reactor heat flux and enthalpy difference)
            self.sheet[0].cell(row=self.counter[3][1]-1, column=1).value = self.process_point
            heat_flux_reactor = Dictionary.calorimeter_thermostat["{:d}".format(int(self.cur_operation_point.temperature))].forward(mean_values[3:])
            heat_flux_reactor.insert(1, heat_flux_reactor[0]-sum(heat_flux_outside[:2]))

            for idx in range(len(heat_flux_reactor)):
                self.sheet[0].cell(row=self.counter[3][1]-1, column=idx+4).value = heat_flux_reactor[idx]

            enthalpy_difference = (sum(heat_flux_reactor[1:])+heat_flux_outside[2]) / (self.actual_molar_flowrate[0]*1000)
            self.sheet[0].cell(row=self.counter[3][1]-1, column=self.counter[3][2]).value = enthalpy_difference

            # save changes 
            self.workbook.save("Calorimetry\{0}.xlsx".format(self.excel_name))

    def point_complete(self):
        if self.state == Output_Calculation_Absolute_Evaluation.States.TEMPERATURE_EQUILIBRATION:
            val = 10
            if len(self.datalist) < val:
                return False
        
            for col_idx in [2, 3, 4]:
                valid_count = 0
                for idx in range(val):
                    if abs(self.datalist[len(self.datalist)-1-idx][col_idx] - self.cur_operation_point.get_temperature()) < 0.1:
                        valid_count += 1
                if valid_count < math.ceil(val*0.9):
                    # return False
                    return True
            return True
        elif self.state == Output_Calculation_Absolute_Evaluation.States.SETTING_DEADLINE:
            self.cur_deadline = time.monotonic_ns() + self.cur_operation_point.get_time_ms() * 1E6
            self.min_time = time.monotonic_ns() +  self.dead_time
            self.state = Output_Calculation_Absolute_Evaluation.States.WAITING_FOR_DEADLINE
            self.waiting_counter = 0
            return False
        elif self.state == Output_Calculation_Absolute_Evaluation.States.WAITING_FOR_DEADLINE:
            if self.cur_deadline < time.monotonic_ns():
                self.workbook.save("Calorimetry\{0}.xlsx".format(self.excel_name))
                return True
            else:
                return False
        raise Exception("You should not land here")

    def has_error(self):
        if not self.state == Output_Calculation_Absolute_Evaluation.States.TEMPERATURE_EQUILIBRATION:
            return False

        if self.cur_deadline < time.monotonic_ns():
            print("set_temp is not reached at the reactor")
            return True
        return False

    def push_actual_flowrate(self, val):
        self.actual_flowrate_list = val

    def get_finish_instruction(self):
        # generate charts
        Dia_Raw_Temp = LineChart()

        Dia_Raw_Temp.y_axis.title = "Temperature [°C]"
        y_data = Reference(self.sheet[1], min_col = 2, min_row = 1, max_col = 8, max_row = len(self.datalist)+1)
        Dia_Raw_Temp.add_data(y_data, titles_from_data = True)

        Dia_Raw_Temp.x_axis.title = "Time [s]"
        Dia_Raw_Temp.x_axis.tickLblSkip = math.ceil(len(self.datalist)/10)
        x_data = Reference(self.sheet[1], min_col = 1, min_row = 2, max_row = len(self.datalist)+1)
        Dia_Raw_Temp.set_categories(x_data)

        chart1 = self.workbook.create_chartsheet("Dia_Raw_Temp")
        chart1.add_chart(Dia_Raw_Temp)

        Dia_Raw_Voltage = LineChart()

        Dia_Raw_Voltage.y_axis.title = "Voltage [mV]"
        y_data = Reference(self.sheet[1], min_col = 9, min_row = 1, max_col = 11, max_row = len(self.datalist)+1)
        Dia_Raw_Voltage.add_data(y_data, titles_from_data = True)

        Dia_Raw_Voltage.x_axis.title = "Time [s]"
        Dia_Raw_Voltage.x_axis.tickLblSkip = math.ceil(len(self.datalist)/10)
        x_data = Reference(self.sheet[1], min_col = 1, min_row = 2, max_row = len(self.datalist)+1)
        Dia_Raw_Voltage.set_categories(x_data)

        chart2 = self.workbook.create_chartsheet("Dia_Raw_Voltage")
        chart2.add_chart(Dia_Raw_Voltage)

        # formatting
        substance_a_color = "3BCCFF"
        substance_b_color = "3D33FF"
        result_color = "FF087F"
        add_data_color = "4B0082"

        for idx in range(4):
            self.sheet[0].cell(row=self.counter[idx][0], column=1).font = Font(bold=True)
            self.sheet[0].cell(row=self.counter[idx][0], column=1).alignment = Alignment(horizontal="center")
            self.sheet[0].merge_cells(start_row=self.counter[idx][0], start_column=1, end_row=self.counter[idx][0], end_column=self.counter[idx][2])
        
            for jdx in range(1,self.counter[idx][2]+1):
                self.sheet[0].cell(row=self.counter[idx][0]+1, column=jdx).border = Border(bottom=Side(border_style="thick"))
        
        for idx in range(self.process_point):
            self.sheet[0].cell(row=self.counter[1][0]+idx+2, column=5).fill = PatternFill("lightUp", fgColor=substance_a_color)
            self.sheet[0].cell(row=self.counter[1][0]+idx+2, column=9).fill = PatternFill("lightUp", fgColor=substance_b_color)
            self.sheet[0].cell(row=self.counter[3][0]+idx+2, column=8).fill = PatternFill("lightUp", fgColor=result_color)
        
        for idx in range(1,self.counter[0][2]+1):
            self.sheet[0].cell(row=3, column=idx).fill = PatternFill("lightTrellis", fgColor=substance_a_color)
            self.sheet[0].cell(row=4, column=idx).fill = PatternFill("lightTrellis", fgColor=substance_b_color)
                
        self.sheet[0].cell(row=2, column=8).font = Font(bold=True)
        self.sheet[0].cell(row=2, column=8).alignment = Alignment(horizontal="center")
        self.sheet[0].merge_cells(start_row=2, start_column=8, end_row=2, end_column=9)
        for idx in range(2):
            self.sheet[0].cell(row=idx+3, column=9).fill = PatternFill("lightTrellis", fgColor=add_data_color)
            self.sheet[0].cell(row=3, column=idx+8).border = Border(top=Side(border_style="thick"))    
        
        self.workbook.save("Calorimetry\{0}.xlsx".format(self.excel_name))
